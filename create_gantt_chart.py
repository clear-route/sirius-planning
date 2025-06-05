import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import traceback

# Hardcoded activity data to bypass markdown parsing issues
ACTIVITIES_DATA = [
    {"type": "engineer_header", "name": "Engineer 1: Central UAT Test Case Identification & Migration"},
    {"type": "phase_header", "name": "Phase 1: Discovery, Analysis & Planning (Est. Months 1-2)"},
    {"type": "activity", "activity_num": "1.", "name": "Deep Dive into Existing UAT Processes & Test Assets", "timeline_raw_str": "Weeks 1-3 (~10-12 person-days)"},
    {"type": "activity", "activity_num": "2.", "name": "Identify & Prioritize UAT Scenarios for Automation", "timeline_raw_str": "Weeks 3-6 (~12-15 person-days, overlapping with activity 1 & 3)"},
    {"type": "activity", "activity_num": "3.", "name": "Master BDD Tooling & Methodology", "timeline_raw_str": "Weeks 2-5 (Can be concurrent; dedicated learning ~5-7 person-days)"},
    {"type": "phase_header", "name": "Phase 2: Migration, Automation Development & Initial Integration (Est. Months 3-4)"},
    {"type": "activity", "activity_num": "4.", "name": "Convert Selected UAT Scenarios to BDD (Gherkin)", "timeline_raw_str": "Weeks 7-10 (~15-18 person-days)"},
    {"type": "activity", "activity_num": "5.", "name": "Develop Automated Test Scripts using Playwright", "timeline_raw_str": "Weeks 9-16 (~25-30 person-days, significant overlap with activity 4 initially, then focused development)"},
    {"type": "activity", "activity_num": "6.", "name": "Setup & Test Execution in DT2 Environment", "timeline_raw_str": "Weeks 15-18 (~8-10 person-days, concurrent with later stages of Activity 5)"},
    {"type": "phase_header", "name": "Phase 3: Refinement, Reporting & Knowledge Transfer Preparation (Est. Months 5-6)"},
    {"type": "activity", "activity_num": "7.", "name": "Iterate and Refine Automated UAT Suite", "timeline_raw_str": "Weeks 17-24 (~15-20 person-days, ongoing)"},
    {"type": "activity", "activity_num": "8.", "name": "Establish Automated UAT Reporting", "timeline_raw_str": "Weeks 19-22 (~7-10 person-days)"},
    {"type": "activity", "activity_num": "9.", "name": "Document Best Practices & Create Migration Playbook", "timeline_raw_str": "Weeks 20-24 (~8-10 person-days, ongoing documentation build-up)"},
    {"type": "activity", "activity_num": "10.", "name": "Prepare for Knowledge Sharing & Team Onboarding", "timeline_raw_str": "Weeks 22-24 (~5-7 person-days)"},

    {"type": "engineer_header", "name": "Engineer 2: Embedding New Ways of Working & Engineering Practices"},
    {"type": "phase_header", "name": "Phase 1: Assessment, Strategy Definition & Foundational Setup (Est. Months 1-2)"},
    {"type": "activity", "activity_num": "1.", "name": "Baseline Current Engineering Practices & CI/CD Maturity", "timeline_raw_str": "Weeks 1-3 (~10-12 person-days)"},
    {"type": "activity", "activity_num": "2.", "name": "Develop & Communicate Pilot Engineering Practices Adoption Strategy", "timeline_raw_str": "Weeks 2-4 (~7-8 person-days)"},
    {"type": "activity", "activity_num": "3.", "name": "Tooling Onboarding & Environment Preparation", "timeline_raw_str": "Weeks 3-6 (Can be concurrent; dedicated effort ~8-10 person-days)"},
    {"type": "phase_header", "name": "Phase 2: Implementation, Coaching & CI/CD Integration (Est. Months 3-4)"},
    {"type": "activity", "activity_num": "4.", "name": "Drive Adoption of Unit Testing & Developer-Led Testing", "timeline_raw_str": "Weeks 7-16 (~20-25 person-days, ongoing coaching and support)"},
    {"type": "activity", "activity_num": "5.", "name": "Integrate Automated Tests into CI/CD Pipelines (GitHub Actions Focus)", "timeline_raw_str": "Weeks 9-16 (~20-25 person-days, heavy collaboration with squads)"},
    {"type": "activity", "activity_num": "6.", "name": "Establish & Champion Mocking Practices (Mockito/MockFlow)", "timeline_raw_str": "Weeks 10-15 (~10-12 person-days, concurrent with other activities)"},
    {"type": "phase_header", "name": "Phase 3: Optimization, Standardization & Knowledge Dissemination (Est. Months 5-6)"},
    {"type": "activity", "activity_num": "7.", "name": "Refine CI/CD Pipelines (GitHub Actions) and Test Execution Efficiency", "timeline_raw_str": "Weeks 17-24 (~10-15 person-days, ongoing)"},
    {"type": "activity", "activity_num": "8.", "name": "Develop & Document Standardized Engineering Playbooks", "timeline_raw_str": "Weeks 18-24 (~15-18 person-days, ongoing documentation build-up)"},
    {"type": "activity", "activity_num": "9.", "name": "Facilitate Performance Profiling Setup", "timeline_raw_str": "Weeks 20-23 (~5-7 person-days)"},
    {"type": "activity", "activity_num": "10.", "name": "Prepare for Scaling & Knowledge Transfer", "timeline_raw_str": "Weeks 22-24 (~7-10 person-days)"}
]

def extract_week_str(timeline_raw_str):
    if not timeline_raw_str: return ""
    week_match = re.search(r"Weeks\s*(\d+)-(\d+)", timeline_raw_str, re.IGNORECASE)
    if week_match:
        start_w = week_match.group(1)
        end_w = week_match.group(2)
        return f"(W{start_w}-W{end_w})"
    return ""

def parse_timeline_to_sprint_indices(timeline_raw_str, num_total_sprints=12):
    sprint_indices = set()
    if not timeline_raw_str: 
        return []
    
    week_match = re.search(r"Weeks\s*(\d+)-(\d+)", timeline_raw_str, re.IGNORECASE)
    if week_match:
        try:
            start_week = int(week_match.group(1))
            end_week = int(week_match.group(2))
            
            start_sprint = (start_week - 1) // 2 
            end_sprint = (end_week - 1) // 2   
            
            for sprint_idx in range(start_sprint, end_sprint + 1):
                if 0 <= sprint_idx < num_total_sprints:
                    sprint_indices.add(sprint_idx)
            return sorted(list(sprint_indices))
        except ValueError:
            print(f"DEBUG_SPRINT_PARSE_ERROR: Could not parse week numbers from timeline string: {timeline_raw_str}")
            return []
    return sorted(list(sprint_indices))

def create_gantt_excel(data, filename="pilot_gantt_chart_sprints.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pilot Gantt (Sprints)"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    activity_fill = PatternFill(start_color="AEAAAA", end_color="AEAAAA", fill_type="solid")
    engineer_header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    phase_header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid") 
    phase_font = Font(bold=True, size=12, color="FFFFFF")

    num_sprints = 12 
    sprint_headers = [f"Sprint {i+1}\n(W{2*i+1}-W{2*i+2})" for i in range(num_sprints)]
    
    ws.cell(row=1, column=1, value="Activity / Task (Est. Timeline)").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.column_dimensions[get_column_letter(1)].width = 105

    for col_num, sprint_name in enumerate(sprint_headers, 2):
        cell = ws.cell(row=1, column=col_num, value=sprint_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_num)].width = 12

    current_row = 2
    activity_display_count = 0
    for item in data: # Use the hardcoded data directly
        if item["type"] == "engineer_header":
            cell = ws.cell(row=current_row, column=1, value=item["name"])
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.fill = engineer_header_fill
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sprint_headers)+1)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[current_row].height = 22
            current_row += 1
        elif item["type"] == "phase_header":
            cell = ws.cell(row=current_row, column=1, value=item["name"])
            cell.font = phase_font
            cell.fill = phase_header_fill
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sprint_headers)+1)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[current_row].height = 20
            current_row += 1
        elif item["type"] == "activity":
            activity_display_count += 1
            week_str_display = extract_week_str(item.get("timeline_raw_str", ""))
            activity_display_name = f"{item['activity_num']} {item['name']} {week_str_display}".strip()
            name_cell = ws.cell(row=current_row, column=1, value=activity_display_name)
            name_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=2) 
            
            sprint_indices_to_color = parse_timeline_to_sprint_indices(item.get("timeline_raw_str", ""), num_sprints)
            if not item.get("timeline_raw_str") or not sprint_indices_to_color:
                print(f"OUTPUT_WARNING: Activity '{activity_display_name}' ({item['activity_num'][:-1]}) will have no bar (timeline missing or unparsable: '{item.get('timeline_raw_str', 'N/A')}').")
            
            for sprint_idx in sprint_indices_to_color:
                col_to_color = sprint_idx + 2 
                ws.cell(row=current_row, column=col_to_color).fill = activity_fill
            current_row += 1
    
    ws.row_dimensions[1].height = 45 
    ws.freeze_panes = 'B2'

    try:
        wb.save(filename)
        print(f"EXCEL_INFO: Gantt chart '{filename}' saved. Rows for activities written: {activity_display_count}")
    except Exception as e:
        print(f"EXCEL_ERROR: Error saving Excel file '{filename}': {e}")
        traceback.print_exc()
    return activity_display_count 

if __name__ == "__main__":
    print(f"SCRIPT_INFO: Main script execution started using hardcoded data.")
    activities_written_to_excel = 0
    try:
        actual_activities_from_hardcoded_data = [item for item in ACTIVITIES_DATA if item['type'] == 'activity']
        print(f"SCRIPT_INFO: Using {len(actual_activities_from_hardcoded_data)} hardcoded activity entries from ACTIVITIES_DATA list.")
        
        activities_with_timelines = [act for act in actual_activities_from_hardcoded_data if act.get('timeline_raw_str')]
        print(f"SCRIPT_INFO: Out of {len(actual_activities_from_hardcoded_data)} hardcoded activities, {len(activities_with_timelines)} have a timeline_raw_str.")
        if len(activities_with_timelines) < len(actual_activities_from_hardcoded_data):
            missing_timeline_activities = [act['name'] for act in actual_activities_from_hardcoded_data if not act.get('timeline_raw_str')]
            print(f"SCRIPT_WARNING: {len(actual_activities_from_hardcoded_data) - len(activities_with_timelines)} hardcoded activities are missing timeline strings (their bars will be empty):")
            for missing_act_name in missing_timeline_activities:
                print(f"    - Missing timeline for: {missing_act_name}")
        else:
            print("SCRIPT_INFO: All hardcoded activities appear to have timeline strings.")
        
        activities_written_to_excel = create_gantt_excel(ACTIVITIES_DATA, filename="pilot_gantt_chart_sprints.xlsx")
        if activities_written_to_excel > 0:
            print(f"SCRIPT_INFO: Confirmed {activities_written_to_excel} activity rows were written to Excel.")
        elif len(actual_activities_from_hardcoded_data) > 0:
             print(f"SCRIPT_WARNING: Processed {len(actual_activities_from_hardcoded_data)} hardcoded activities but wrote 0 activity display rows to Excel. Check create_gantt_excel item loop.")

    except Exception as e:
        print(f"SCRIPT_CRITICAL_ERROR: An unexpected error occurred during script execution: {e}")
        traceback.print_exc()
